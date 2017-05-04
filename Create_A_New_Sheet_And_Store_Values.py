#!/usr/bin/python3

#import the opepyxl
import openpyxl

#Create a Workbbok object
objWorkbook = openpyxl.Workbook()

#Create a new sheet at the beginning
objWorkbook.create_sheet(index=0,title='First_Sheet')

#Activate the new sheet
#By creating a sheet object
#pointing to it
objSheet = objWorkbook.get_sheet_by_name('First_Sheet')

#Insert values into the sheet
for iRow in range(1,10):
    for iColumn in range(1,7):
        objSheet.cell(row = iRow,column = iColumn).value = iRow

#delete the sheet 'Sheet'
#the argument of the Workbook object
#accepts sheet object
objWorkbook.remove_sheet(objWorkbook.get_sheet_by_name('Sheet'))

#Save the Excel
objWorkbook.save('Create_A_New_Sheet_And_Store_Values.xlsx')
