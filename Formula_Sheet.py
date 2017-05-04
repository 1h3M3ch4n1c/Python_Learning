#!/usr/bin/python3
#import openpyxl
import openpyxl

#create wokrbook object
objWorkbook = openpyxl.Workbook()

#create a newsheet
objWorkbook.create_sheet(index=0,title='Sheet 1')

#create a sheet object
objSheet = objWorkbook.get_sheet_by_name('Sheet 1')

#enter value in the cell A1:A7
for iItr in range(1,8):
    objSheet.cell(row=iItr,column=1).value = iItr*10

#get the sum in A8
objSheet.cell(row=8,column=1).value = '=sum(A1:A7)'

#save the sheet
objWorkbook.save('Formula_Sheet.xlsx')
