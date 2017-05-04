#!/usr/bin/python3
#import openpyxl module
import openpyxl
#import Style and Font
from openpyxl.styles import Font,Style


#create a workbook object
objWorkbook = openpyxl.Workbook()

#create a sheet
objWorkbook.create_sheet(index=0,title='Sheet 1')

#create a sheet object
objSheet = objWorkbook.get_sheet_by_name('Sheet 1')

#Font
italic32Font = Font(name='TimesNewRoman',size=32,bold=True,italic=True)

#Style
objStyle = Style(font = italic32Font)

#Apply the style in cell(A1)
objSheet.cell(row=1,column=1).style = objStyle
objSheet.cell(row=1,column=1).value = 'Font And Style'

#Save the workbook
objWorkbook.save('Font_And_Styling.xlsx')
